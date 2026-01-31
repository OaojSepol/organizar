import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Inventário Filial 944", page_icon="📝")

st.title("📝 Gerador de Inventário - Filial 944")
st.write("Organização automática por abas (incluindo separação de Scaner de Mão via Complemento).")

uploaded_file = st.file_uploader("Escolha o arquivo CSV", type="csv")

if uploaded_file is not None:
    try:
        # Lendo o arquivo original
        df = pd.read_csv(uploaded_file, sep=';')
        
        # --- LÓGICA DE DEFINIÇÃO DAS ABAS (REVISADA) ---
        def definir_aba(linha):
            tipo = str(linha['TIPO']).upper()
            sub_tipo = str(linha.get('SUB TIPO', '')).upper()
            complemento = str(linha.get('COMPLEMENTO', '')).upper()
            
            # 1. Regra para Scanners de Mão (Verifica SUB TIPO ou COMPLEMENTO)
            if tipo == 'SCANER' and ('MÃO' in sub_tipo or 'MÃO' in complemento):
                return 'SCANER DE MÃO'
            
            # 2. Regra para a aba SERVIDOR (Unificada: Servidor, Tape, Rack, Storage)
            tipos_servidor = ['SERVIDOR', 'TAPE', 'RACK', 'STORAGE']
            if tipo in tipos_servidor:
                return 'SERVIDOR'
            
            # 3. Padrão: Usa o próprio TIPO (ex: MONITOR, CPU)
            return tipo

        # Criar a coluna que define o nome da aba
        df['ABA_DESTINO'] = df.apply(definir_aba, axis=1)
        
        # Colunas que serão removidas da visualização final (conforme pedido)
        colunas_remover = ['FILIAL', 'TIPO', 'SUB TIPO', 'COMPLEMENTO', 'ABA_DESTINO']

        if st.button("🚀 Gerar Planilha Organizada"):
            output = BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Pegar nomes das abas únicos e ordenados
                abas = sorted(df['ABA_DESTINO'].unique())
                
                for nome_aba in abas:
                    grupo = df[df['ABA_DESTINO'] == nome_aba]
                    
                    # Ordenar por PIP para ficar organizado
                    grupo_ordenado = grupo.sort_values(by=['PIP'], ascending=True)
                    
                    nome_final_aba = str(nome_aba)[:31].replace('/', '-')
                    tabela_final = grupo_ordenado.drop(columns=colunas_remover, errors='ignore')
                    
                    # Escrever tabela na linha 2 (deixando a linha 1 para o título)
                    tabela_final.to_excel(writer, sheet_name=nome_final_aba, index=False, startrow=1)
                    
                    ws = writer.sheets[nome_final_aba]
                    
                    # --- 1. Título Superior ---
                    titulo_texto = f"inventario filial 944 - {nome_final_aba}"
                    ws.cell(row=1, column=1).value = titulo_texto
                    num_colunas = len(tabela_final.columns)
                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_colunas)
                    
                    # Estilo do Título
                    titulo_cell = ws.cell(row=1, column=1)
                    titulo_cell.font = Font(size=12, bold=True)
                    titulo_cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                    titulo_cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # --- 2. Estilo do Cabeçalho da Tabela (Linha 2) ---
                    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                    header_font = Font(color="FFFFFF", bold=True)
                    for cell in ws[2]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center")

                    # --- 3. Ajuste Automático de Largura das Colunas ---
                    for i, col_name in enumerate(tabela_final.columns, 1):
                        column_letter = get_column_letter(i)
                        
                        # Tamanho inicial baseado no cabeçalho
                        max_length = len(str(col_name))
                        
                        # Verifica todas as linhas daquela coluna
                        for row in range(2, ws.max_row + 1):
                            val = ws.cell(row=row, column=i).value
                            if val:
                                length = len(str(val))
                                if length > max_length:
                                    max_length = length
                        
                        # Define a largura com uma pequena margem
                        ws.column_dimensions[column_letter].width = max_length + 4

            # Preparar download
            st.download_button(
                label="📥 Baixar Inventário Excel",
                data=output.getvalue(),
                file_name="Inventario_Filial_944_Atualizado.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            st.success(f"Arquivo gerado! Identificamos {len(abas)} categorias diferentes.")

    except Exception as e:
        st.error(f"Erro ao processar: {e}")
